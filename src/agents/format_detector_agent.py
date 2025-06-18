"""Simple script to test format detection using semantic similarity.

Run:
    python simple_detection.py \
        --formats db.xlsx \
        --file test.xlsx
"""

import argparse
import json
from pathlib import Path
from typing import Dict, List, Tuple, Set

import pandas as pd
import numpy as np
from sentence_transformers import SentenceTransformer
from openpyxl import load_workbook
from openpyxl.cell import MergedCell


# Global configuration
COLUMN_MATCH_SIMILARITY_THRESHOLD = 0.5  # Lowered threshold to catch more variations
DEBUG_MODE = True  # Set to True to see detailed matching information


class SimpleFormatDetector:
    """Minimal format detector using semantic similarity."""
    
    def __init__(self):
        """Initialize the sentence transformer model."""
        self.model = SentenceTransformer('all-MiniLM-L6-v2')
        self.known_formats = {}
        
    def load_formats(self, excel_path: Path) -> None:
        """Load known formats from Excel file."""
        df = pd.read_excel(excel_path)
        
        for _, row in df.iterrows():
            format_label = row.get('Format Label', '')
            if not format_label:
                continue
                
            # Parse sample columns from the layout
            sample_cols = []
            layout = row.get('Sample Layout', '')
            
            if isinstance(layout, str) and layout.strip():
                try:
                    # Try to parse as Python literal
                    import ast
                    parsed = ast.literal_eval(layout)
                    if isinstance(parsed, list) and parsed:
                        sample_cols = list(parsed[0].keys())
                except:
                    # If parsing fails, try to extract column names manually
                    pass
            
            if sample_cols:
                # Clean column names
                sample_cols = [str(col).strip() for col in sample_cols]
                
                self.known_formats[format_label] = {
                    'columns': sample_cols,
                    'description': row.get('Description', ''),
                    'category': row.get('Format Type', 'Unknown')
                }
        
        print(f"Loaded {len(self.known_formats)} formats from {excel_path}")
        
        if DEBUG_MODE and self.known_formats:
            print("\n🔍 DEBUG - Sample of loaded formats:")
            for fmt_name, fmt_info in list(self.known_formats.items())[:2]:
                print(f"  Format: {fmt_name}")
                print(f"  Columns: {fmt_info['columns'][:5]}...")
                if len(fmt_info['columns']) > 5:
                    print(f"  ... and {len(fmt_info['columns']) - 5} more columns")
    
    def handle_merged_cells(self, file_path: Path, sheet_name: str, max_rows: int = 10) -> List[str]:
        """Extract headers from Excel sheet handling merged cells."""
        wb = load_workbook(file_path, data_only=True)
        ws = wb[sheet_name]
        
        # Find all merged cell ranges
        merged_cells = list(ws.merged_cells.ranges)
        
        if DEBUG_MODE and merged_cells:
            print(f"   Found {len(merged_cells)} merged cell ranges in sheet '{sheet_name}'")
        
        # Process first few rows to find headers
        potential_headers = []
        for row_idx in range(1, min(max_rows + 1, ws.max_row + 1)):
            row_values = []
            
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                
                # Check if this cell is part of a merged range
                cell_value = None
                if isinstance(cell, MergedCell):
                    # Find the merged range this cell belongs to
                    for merged_range in merged_cells:
                        if cell.coordinate in merged_range:
                            # Get value from the top-left cell of the range
                            top_left = ws.cell(merged_range.min_row, merged_range.min_col)
                            cell_value = top_left.value
                            break
                else:
                    cell_value = cell.value
                
                # Clean the value
                if cell_value is None:
                    cell_value = ""
                else:
                    cell_value = str(cell_value).strip()
                
                row_values.append(cell_value)
            
            # Check if this row could be headers
            non_empty = [v for v in row_values if v and v.lower() not in ['', 'nan', 'none']]
            if len(non_empty) >= len(row_values) * 0.3:  # At least 30% non-empty
                potential_headers.append((row_idx - 1, row_values))
        
        wb.close()
        
        # Find the best header row
        if potential_headers:
            # Prefer rows with more non-empty values
            best_row = max(potential_headers, key=lambda x: len([v for v in x[1] if v]))
            if DEBUG_MODE:
                print(f"   Selected headers from row {best_row[0] + 1}: {best_row[1][:5]}...")
            return best_row[1], best_row[0]
        
        return None, None
        """Load known formats from Excel file."""
        df = pd.read_excel(excel_path)
        
        for _, row in df.iterrows():
            format_label = row.get('Format Label', '')
            if not format_label:
                continue
                
            # Parse sample columns from the layout
            sample_cols = []
            layout = row.get('Sample Layout', '')
            
            if isinstance(layout, str) and layout.strip():
                try:
                    # Try to parse as Python literal
                    import ast
                    parsed = ast.literal_eval(layout)
                    if isinstance(parsed, list) and parsed:
                        sample_cols = list(parsed[0].keys())
                except:
                    # If parsing fails, try to extract column names manually
                    pass
            
            if sample_cols:
                # Clean column names
                sample_cols = [str(col).strip() for col in sample_cols]
                
                self.known_formats[format_label] = {
                    'columns': sample_cols,
                    'description': row.get('Description', ''),
                    'category': row.get('Format Type', 'Unknown')
                }
        
        print(f"Loaded {len(self.known_formats)} formats from {excel_path}")
        
        if DEBUG_MODE and self.known_formats:
            print("\n🔍 DEBUG - Sample of loaded formats:")
            for fmt_name, fmt_info in list(self.known_formats.items())[:2]:
                print(f"  Format: {fmt_name}")
                print(f"  Columns: {fmt_info['columns'][:5]}...")
                if len(fmt_info['columns']) > 5:
                    print(f"  ... and {len(fmt_info['columns']) - 5} more columns")
        
    def compute_column_match_score(self, file_cols: List[str], format_cols: List[str]) -> Dict:
        """Compute multiple metrics for column matching."""
        # Compute embeddings for all columns
        if not file_cols or not format_cols:
            return {
                'exact_match_ratio': 0,
                'coverage': 0,
                'precision': 0,
                'semantic_similarity': 0,
                'order_score': 0,
                'composite_score': 0,
                'exact_matches': 0,
                'semantic_matches': 0,
                'total_format_cols': len(format_cols),
                'total_file_cols': len(file_cols),
                'match_details': []
            }
        
        # Get embeddings for all columns
        file_embeddings = self.model.encode(file_cols)
        format_embeddings = self.model.encode(format_cols)
        
        # Find best matches for each format column
        match_details = []
        matched_file_indices = set()
        
        for i, format_col in enumerate(format_cols):
            best_match = None
            best_score = 0
            best_file_idx = -1
            
            for j, file_col in enumerate(file_cols):
                # Skip if already matched
                if j in matched_file_indices:
                    continue
                
                # Compute similarity
                similarity = float(np.dot(format_embeddings[i], file_embeddings[j]) / (
                    np.linalg.norm(format_embeddings[i]) * np.linalg.norm(file_embeddings[j])
                ))
                
                if similarity > best_score:
                    best_score = similarity
                    best_match = file_col
                    best_file_idx = j
            
            # Consider it a match if similarity > threshold
            if best_score > COLUMN_MATCH_SIMILARITY_THRESHOLD and best_file_idx != -1:
                matched_file_indices.add(best_file_idx)
                match_details.append({
                    'format_col': format_col,
                    'matched_col': best_match,
                    'similarity': best_score,
                    'is_exact': format_col.lower().strip() == best_match.lower().strip()
                })
            elif DEBUG_MODE and best_score > 0.3:  # Show near-misses in debug mode
                print(f"     Near-miss: '{format_col}' ~~ '{best_match}' (similarity: {best_score:.3f})")
                matched_file_indices.add(best_file_idx)
                match_details.append({
                    'format_col': format_col,
                    'matched_col': best_match,
                    'similarity': best_score,
                    'is_exact': format_col.lower().strip() == best_match.lower().strip()
                })
        
        # Calculate metrics
        total_matches = len(match_details)
        exact_matches = sum(1 for m in match_details if m['is_exact'])
        semantic_matches = total_matches - exact_matches
        
        # 1. Match ratio (semantic + exact)
        match_ratio = total_matches / len(format_cols) if format_cols else 0
        
        # 2. Coverage (how many required columns are present semantically)
        coverage = total_matches / len(format_cols) if format_cols else 0
        
        # 3. Precision (how many file columns are relevant)
        precision = len(matched_file_indices) / len(file_cols) if file_cols else 0
        
        # 4. Overall semantic similarity
        if file_cols and format_cols:
            text1 = ' '.join(file_cols)
            text2 = ' '.join(format_cols)
            embeddings = self.model.encode([text1, text2])
            semantic_similarity = float(np.dot(embeddings[0], embeddings[1]) / (
                np.linalg.norm(embeddings[0]) * np.linalg.norm(embeddings[1])
            ))
        else:
            semantic_similarity = 0.0
        
        # 5. Column order similarity (bonus if columns are in similar order)
        order_score = 0.0
        if total_matches > 0:
            matched_positions = []
            for match in match_details:
                try:
                    file_pos = file_cols.index(match['matched_col'])
                    format_pos = format_cols.index(match['format_col'])
                    matched_positions.append((format_pos, file_pos))
                except ValueError:
                    pass
            
            if len(matched_positions) >= 2:
                # Check if relative ordering is preserved
                order_preserved = 0
                for i in range(len(matched_positions) - 1):
                    if matched_positions[i][1] < matched_positions[i+1][1]:
                        order_preserved += 1
                order_score = order_preserved / (len(matched_positions) - 1)
        
        # 6. Average match quality
        avg_match_quality = sum(m['similarity'] for m in match_details) / len(match_details) if match_details else 0
        
        # 7. Composite score (weighted combination)
        composite_score = (
            0.25 * match_ratio +           # Overall match rate
            0.20 * coverage +              # Required columns covered
            0.15 * precision +             # File columns that are relevant
            0.20 * avg_match_quality +     # Quality of matches
            0.15 * semantic_similarity +   # Overall semantic similarity
            0.05 * order_score            # Column order preservation
        )
        
        return {
            'match_ratio': match_ratio,
            'coverage': coverage,
            'precision': precision,
            'semantic_similarity': semantic_similarity,
            'order_score': order_score,
            'avg_match_quality': avg_match_quality,
            'composite_score': composite_score,
            'exact_matches': exact_matches,
            'semantic_matches': semantic_matches,
            'total_matches': total_matches,
            'total_format_cols': len(format_cols),
            'total_file_cols': len(file_cols),
            'match_details': match_details
        }
    
    def detect_format_in_sheet(self, df: pd.DataFrame, sheet_name: str, file_path: Path = None) -> List[Dict]:
        """Detect format in a single sheet."""
        file_columns = list(df.columns)
        
        # Skip sheets with very few columns or rows
        if len(file_columns) < 2 or len(df) < 1:
            return []
        
        # Check if we have unnamed columns (common when headers aren't in row 1)
        unnamed_count = sum(1 for col in file_columns if str(col).startswith('Unnamed:'))
        
        if unnamed_count > len(file_columns) * 0.5 and file_path:  # More than 50% unnamed
            if DEBUG_MODE:
                print(f"\n⚠️  Sheet '{sheet_name}' has {unnamed_count}/{len(file_columns)} unnamed columns")
                print("   Checking for merged cells and actual headers...")
            
            # Try to handle merged cells
            better_headers, header_row_idx = self.handle_merged_cells(file_path, sheet_name)
            
            if better_headers:
                # Use the better headers we found
                file_columns = better_headers[:len(df.columns)]  # Ensure same length
                if DEBUG_MODE:
                    print(f"   ✓ Found better headers from merged cells: {file_columns[:5]}...")
            else:
                # Fallback: Try to find headers in the dataframe itself
                actual_headers_found = False
                for idx in range(min(5, len(df))):
                    potential_headers = df.iloc[idx].astype(str).tolist()
                    # Check if this row has meaningful headers
                    non_empty = [h for h in potential_headers if h and not h.startswith('nan')]
                    if len(non_empty) >= len(potential_headers) * 0.5:  # At least 50% non-empty
                        file_columns = potential_headers
                        actual_headers_found = True
                        if DEBUG_MODE:
                            print(f"   ✓ Found potential headers in row {idx + 1}: {file_columns[:5]}...")
                        break
                
                if not actual_headers_found and DEBUG_MODE:
                    print("   ✗ No clear headers found, using original column names")
        
        # Clean column names (remove extra spaces, convert to string)
        file_columns = [str(col).strip() for col in file_columns]
        
        # Additional cleaning for merged cell artifacts
        file_columns = [col if col and col.lower() not in ['', 'nan', 'none'] 
                       else f'Column_{i}' for i, col in enumerate(file_columns)]
        
        if DEBUG_MODE and sheet_name not in ['RFP Info']:  # Debug specific sheets
            print(f"\n🔍 DEBUG - Sheet '{sheet_name}' final columns:")
            print(f"   File columns ({len(file_columns)}): {file_columns[:10]}")
            if len(file_columns) > 10:
                print(f"   ... and {len(file_columns) - 10} more")
        
        results = []
        for format_name, format_info in self.known_formats.items():
            scores = self.compute_column_match_score(
                file_columns, 
                format_info['columns']
            )
            
            # Debug output for low scores
            if DEBUG_MODE and scores['composite_score'] < 0.1 and len(results) < 3:
                print(f"\n🔍 DEBUG - Low score for format '{format_name}':")
                print(f"   Expected columns: {format_info['columns'][:5]}...")
                print(f"   Composite score: {scores['composite_score']:.3f}")
                print(f"   Matches: {scores['total_matches']}/{scores['total_format_cols']}")
                if scores['match_details']:
                    print(f"   Best matches found:")
                    for m in scores['match_details'][:3]:
                        print(f"     - '{m['format_col']}' → '{m['matched_col']}' (sim: {m['similarity']:.3f})")
            
            results.append({
                'format': format_name,
                'sheet': sheet_name,
                'scores': scores,
                'category': format_info['category'],
                'description': format_info['description'],
                'expected_columns': format_info['columns'],
                'detected_columns': file_columns
            })
        
        return results
    
    def detect_format(self, file_path: Path, header_row: int = None) -> Dict:
        """Detect format across all sheets in the file."""
        all_results = []
        
        if file_path.suffix.lower() in {'.xlsx', '.xls'}:
            # Load all sheets
            excel_file = pd.ExcelFile(file_path)
            sheet_names = excel_file.sheet_names
            
            print(f"\nAnalyzing {len(sheet_names)} sheets in {file_path.name}")
            
            for sheet_name in sheet_names:
                if header_row is not None:
                    # Use specified header row
                    df = pd.read_excel(file_path, sheet_name=sheet_name, header=header_row)
                else:
                    # Let pandas detect or we'll find headers manually
                    df = pd.read_excel(file_path, sheet_name=sheet_name)
                
                print(f"  - Sheet '{sheet_name}': {len(df)} rows, {len(df.columns)} columns")
                
                sheet_results = self.detect_format_in_sheet(df, sheet_name, file_path)
                all_results.extend(sheet_results)
        else:
            # CSV has only one "sheet"
            if header_row is not None:
                df = pd.read_csv(file_path, header=header_row)
            else:
                df = pd.read_csv(file_path)
            sheet_results = self.detect_format_in_sheet(df, 'main', file_path)
            all_results.extend(sheet_results)
        
        # Sort by composite score
        all_results.sort(key=lambda x: x['scores']['composite_score'], reverse=True)
        
        # Get best match
        best_match = all_results[0] if all_results else None
        
        # Group results by sheet for summary
        results_by_sheet = {}
        for result in all_results:
            sheet = result['sheet']
            if sheet not in results_by_sheet:
                results_by_sheet[sheet] = []
            results_by_sheet[sheet].append(result)
        
        return {
            'file': str(file_path),
            'sheets_analyzed': len(results_by_sheet),
            'best_match': best_match,
            'top_10_matches': all_results[:10],
            'results_by_sheet': results_by_sheet
        }


def main():
    """Main function."""
    parser = argparse.ArgumentParser(description='Simple format detection')
    parser.add_argument(
        '--formats',
        type=Path,
        required=True,
        help='Excel file with format definitions'
    )
    parser.add_argument(
        '--file',
        type=Path,
        required=True,
        help='File to detect format'
    )
    parser.add_argument(
        '--header-row',
        type=int,
        default=None,
        help='Row number containing headers (0-based). If not specified, will auto-detect.'
    )
    
    args = parser.parse_args()
    
    # Initialize detector
    detector = SimpleFormatDetector()
    
    # Load known formats
    detector.load_formats(args.formats)
    
    # Detect format
    result = detector.detect_format(args.file, args.header_row)
    
    # Print results
    print("\n" + "="*60)
    print(f"Format Detection Results for: {result['file']}")
    print("="*60)
    
    if result['best_match']:
        best = result['best_match']
        scores = best['scores']
        print(f"\n🏆 BEST MATCH:")
        print(f"  Format: {best['format']}")
        print(f"  Sheet: {best['sheet']}")
        print(f"  Category: {best['category']}")
        print(f"\n  Scores:")
        print(f"    - Composite Score: {scores['composite_score']:.3f}")
        print(f"    - Match Ratio: {scores['match_ratio']:.3f}")
        print(f"    - Coverage: {scores['coverage']:.3f} ({scores['total_matches']}/{scores['total_format_cols']} columns)")
        print(f"    - Avg Match Quality: {scores['avg_match_quality']:.3f}")
        print(f"    - Exact Matches: {scores['exact_matches']}")
        print(f"    - Semantic Matches: {scores['semantic_matches']}")
        print(f"    - Precision: {scores['precision']:.3f}")
        print(f"    - Order Preservation: {scores['order_score']:.3f}")
        
        # Show column mappings
        if scores.get('match_details'):
            print(f"\n  Column Mappings:")
            for match in scores['match_details'][:5]:  # Show first 5
                exact_label = "✓" if match['is_exact'] else "≈"
                print(f"    {exact_label} '{match['format_col']}' → '{match['matched_col']}' (similarity: {match['similarity']:.3f})")
            if len(scores['match_details']) > 5:
                print(f"    ... and {len(scores['match_details']) - 5} more matches")
        
    print(f"\n📊 TOP 5 MATCHES ACROSS ALL SHEETS:")
    for i, match in enumerate(result['top_10_matches'][:5], 1):
        scores = match['scores']
        print(f"\n  {i}. {match['format']} (Sheet: {match['sheet']})")
        print(f"     Composite Score: {scores['composite_score']:.3f}")
        print(f"     Coverage: {scores['total_matches']}/{scores['total_format_cols']} columns matched")
        print(f"     (Exact: {scores['exact_matches']}, Semantic: {scores['semantic_matches']})")
    
    # Show per-sheet summary
    print(f"\n📋 PER-SHEET SUMMARY:")
    for sheet, sheet_results in result['results_by_sheet'].items():
        best_in_sheet = sheet_results[0]
        print(f"\n  Sheet '{sheet}':")
        print(f"    Best match: {best_in_sheet['format']} (Score: {best_in_sheet['scores']['composite_score']:.3f})")
        print(f"    Columns detected: {len(best_in_sheet['detected_columns'])}")
    
    # Decision threshold
    if result['best_match']:
        confidence = result['best_match']['scores']['composite_score']
        if confidence >= 0.8:
            print(f"\n✅ HIGH CONFIDENCE: Format is likely '{result['best_match']['format']}'")
        elif confidence >= 0.6:
            print(f"\n⚠️  MEDIUM CONFIDENCE: Format might be '{result['best_match']['format']}'")
        else:
            print(f"\n❌ LOW CONFIDENCE: No clear format match found")
    
    # Save full results to JSON
    output_file = Path('detection_results.json')
    with open(output_file, 'w') as f:
        json.dump(result, f, indent=2, default=str)
    print(f"\n💾 Full results saved to: {output_file}")


if __name__ == '__main__':
    main()
